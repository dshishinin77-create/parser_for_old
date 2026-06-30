from ast import Dict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import re
import os
import sys
import copy
import datetime
import json
import traceback


def smart_load_workbook(file_path):
    return load_workbook(file_path)


def get_cell(ini_cell, row, column):
    return (ini_cell[0] + row, ini_cell[1] + column)


def first_cell_def(cells):
    for key in sorted(list(cells.keys())):
        if cells[key]:
            return key


def first_not_empty(ini_cell, cells, direction):
    if ini_cell not in cells.keys():
        return (None, None)
    else:
        if cells[ini_cell]:
            return (ini_cell, cells[ini_cell])
        else:
            if direction == 'row':
                return first_not_empty(get_cell(ini_cell, 0, 1), cells,
                                       direction)
            else:
                return first_not_empty(get_cell(ini_cell, 1, 0), cells,
                                       direction)


def re_var(consts: list, string: str, stop_words=None):
    string = str(string)
    string = ''.join(string.replace('\n', ''))
    reg = '.*' + '.*'.join(consts) + '.*'
    if stop_words:
        for word in stop_words:
            stop = '.*' + word + '.*'
            if bool(re.match(f'{stop}', string, re.IGNORECASE)):
                return False
    return bool(re.match(f'{reg}', string, re.IGNORECASE))


def header_in_reg(header, position, dict_of_reg):
    val = dict_of_reg[position]
    if type(val[0]) == str:
        _temp_bool = re_var(val, header)
    else:
        consts = val[0]
        stop_words = val[1] if len(val) > 1 else None
        _temp_bool = re_var(consts, header, stop_words)
    return _temp_bool


def temp_dict_of_row(row, cells):
    _temp_dict = {}
    if not cells:
        return _temp_dict

    last_cell = list(cells.keys())[-1]
    first_col = None

    for col in range(1, last_cell[1] + 1):
        if cells.get((row, col)):
            first_col = col
            break

    if not first_col:
        return _temp_dict

    cell = (row, first_col)
    value = cells[cell]

    while cell and cell[1] <= last_cell[1]:
        _temp_dict[cell[1]] = value
        cell = (cell[0], cell[1] + 1)
        cell, value = first_not_empty(cell, cells, 'row')
        if not cell or cell[0] != row:
            break
    return _temp_dict


def smart_join(massive):
    massive = [i for i in massive if i]
    return '\n'.join(massive)


# Максимально всеядная функция поиска галочек
def parse_impact_evaluation(cells, CR_d):
    impact_fields = {
        'Detailed_Design': ['detailed', 'design'],
        'Licensing_Documentation': ['licensing', 'documentation'],
        'Structural_Reliability': ['structural', 'reliability'],
        'Industrial_Safety': ['industrial', 'safety'],
        'Ecology_Impact': ['ecology', 'impact'],
        'Fire_Safety': ['fire', 'safety'],
        'Schedule_Impact': ['schedule', 'ütemezés'],
        'Cost_Impact': ['cost', 'impact'],
        'Contract_Impact': ['contract', 'szerződés']
    }

    # Базы символов и слов (включая Marlett и Wingdings)
    checked_exact = ['x', 'v', 'yes', 'igen', '1', '1.0', '+', 'true', 'да',
                     'a', 'r', 'p', 'ü', 'ý', 'þ', '☑', '☒', '✓', '✔', '√',
                     'y', '*']
    unchecked_exact = ['o', '0', '0.0', '-', 'q', 'false', 'no', 'nem', 'нет',
                       '☐', '¨', '£', 'm', 'n/a', 'na', 'n', 'f']
    checked_symbols = ['☑', '☒', '✓', '✔', '√', 'þ', 'ý', 'ü']
    unchecked_symbols = ['☐', '¨']

    for (r, c), val in cells.items():
        if not val or not isinstance(val, str):
            continue

        val_lower = val.strip().lower().replace('\n', ' ')

        for key, keywords in impact_fields.items():
            if all(kw in val_lower for kw in keywords):
                is_checked = False

                # 1. Проверяем внутри самой ячейки текста (если напечатано слитно)
                if any(char in val_lower for char in
                       checked_symbols + ['[x]', '(x)', '[v]']):
                    is_checked = True
                elif any(char in val_lower for char in
                         unchecked_symbols + ['[ ]', '()']):
                    is_checked = False
                else:
                    # 2. Ищем символ в ячейках левее (с запасом на случай объединения)
                    for offset in range(1, 4):
                        if c - offset < 1: break
                        check_val = cells.get((r, c - offset))
                        if check_val is not None:
                            cv_str = str(check_val).strip().lower()
                            if not cv_str:
                                continue

                            # Если текст длинный, это не галочка, а другой столбец — пропускаем
                            if len(cv_str) > 10:
                                continue

                            # Проверяем точные совпадения
                            if cv_str in checked_exact or any(
                                    sym in cv_str for sym in checked_symbols):
                                is_checked = True
                                break
                            elif cv_str in unchecked_exact or any(
                                    sym in cv_str for sym in
                                    unchecked_symbols):
                                is_checked = False
                                break

                            # Умный Fallback: если юзер вбил любой свой короткий символ (1-2 знака)
                            # и это не ноль/буква 'о' из списка выше, считаем, что чекбокс нажат.
                            if len(cv_str) <= 2:
                                is_checked = True
                                break

                if is_checked:
                    CR_d['General_information']['Impact_Evaluation'][
                        key] = 'YES'


dict_of_reg_value_FCR = {
    'CR_number': [['(change.*request|registr|bejegyz)'], ['init']],
    'Reg_date': ['registr', 'dat'],
    'CR_coordinator': ['coord'],
    'CR_number_int': ['init', 'intern'],
    'Organization': [['init', 'organ'], ['intern']],
    'Initiator': [['init'], ['posit', 'organ', 'intern']],
    'Initiator_pos': ['init', 'posit'],
    'Document_type': ['document', 'type'],
    'Change_type': ['change', 'type'],
    'Activity_type': ['activit', 'type'],
    'Constr_facility': [
        ['(object.*construct|construct.*object|építkezés|constr.*facil)'], []],
    'Ini_Method_CR': ['init', 'method', 'cr'],
    'Method_CR': [['method', 'cr'], ['init', 'justif']],
    'Ini_Method_justif': ['init', 'method', 'justif'],
    'Method_justif': [['method', 'justif'], ['init', 'cr']],
    'Change_equipment': ['change', 'equip'],
    'Reason_code': [['(reason.*code|okának.*kód)'], []],
    'CR_reason': [['(reason|okának|other|egyéb)'], ['code', 'kód']],
    'Descr_tech_sol': [['(descr.*sol|descr.*change|módosítások.*leírása)'],
                       []],
    'Final_status': ['final', 'stat'],
    'Material_eq': ['equiv', 'repl'],
    'REPLACE?': ['replac'],
    'Reject_comment': ['reject', 'comment'],
    'Refuse_comment': ['refus', 'comment'],
    'Impact_DDD': ['impact', 'ddd'],
    'Impact_LDD': ['impact', 'ldd'],
    'Impact_cost': ['cost', 'impact'],
    'Schedule': ['schedul'],
    'Prompt_req': ['prompt', 'req'],
    'NS': ['nucl', 'saf'],
    'FS': ['fire', 'saf'],
    'IS': ['industr', 'saf'],
    'ES': ['envir'],
    'SS': ['struct', 'geom']
}

dict_of_reg_value_CRD = {
    'CR_number': [['(change.*request|registr|bejegyz)'], ['init']],
    'CR_coordinator': ['contr', 'coord'],
    'CR_number_int': ['init', 'intern'],
    'Organization': [['init', 'organ'], ['intern']],
    'Initiator': ['init'],
    'Document_type': ['document', 'type'],
    'Method_CR': ['method'],
    'Contract': ['contract'],
    'Impact_cost': ['cost', 'impact'],
    'Schedule': ['schedul'],
    'Comment_nont': ['comment', 'non'],
    'Constr_facility': [['(object.*construct|construct.*object|építkezés)'],
                        []],
    'Reason_code': [['(reason.*code|okának.*kód)'], []],
    'CR_reason': [
        ['(other.*reason|reason.*of.*change|reason.*change|egyéb.*okok)'],
        ['code']],
    'Descr_tech_sol': [['(descr.*sol|descr.*change|módosítások.*leírása)'], []]
}

dict_of_reg_value_CR = {
    'CR_number': [['(change.*request|registr|bejegyz)'], ['init']],
    'Organization': [['init', 'organ'], ['intern']],
    'Responsible': ['responsible', 'eval'],
    'Initiator': [['init'], ['organ', 'intern']],
    'CR_coordinator': ['contr', 'coord'],
    'CR_number_int': ['init', 'intern'],
    'Contract': ['contract'],
    'Impact_cost': ['cost', 'impact'],
    'Schedule': ['schedul'],
    'Comment_nont': ['comment', 'non'],
    'Constr_facility': [['(object.*construct|construct.*object|építkezés)'],
                        []],
    'Reason_code': [['(reason.*code|okának.*kód)'], []],
    'CR_reason': [
        ['(other.*reason|reason.*of.*change|reason.*change|egyéb.*okok)'],
        ['code']],
    'Descr_tech_sol': [['(descr.*sol|descr.*change|módosítások.*leírása)'], []]
}


def main_func(table_name):
    short_table_name = table_name.split('\\')[-1]
    print(f'\n>>> Обработка файла: {short_table_name}')
    warning = 0
    wb = smart_load_workbook(table_name)
    ws = wb.worksheets[0]
    for sheet in wb.worksheets:
        if sheet.sheet_state == 'visible':
            ws = sheet
            break
    cells = ws._cells
    cells = dict(map(lambda x: (x, cells[x].value), cells))
    cells = {
        key: cells[key] if not re_var(['unnamed'], str(cells[key])) else None
        for key in cells}
    last_cell = list(cells.keys())[-1]
    first_cell = first_cell_def(cells)

    text_dump = " ".join([str(cells[k]).lower() for k in cells if cells[k]])

    if 'cr.d' in text_dump or 'fcr.d' in text_dump or 'impact on tdd' in text_dump:
        doc_mode = 'CRD'
        print("    Определен тип: CR.D")
    elif 'field change request' in text_dump or 'fcr' in text_dump:
        doc_mode = 'FCR'
        print("    Определен тип: FCR")
    else:
        doc_mode = 'CR'
        print("    Определен тип: CR")

    if doc_mode == 'CRD':
        dict_of_reg_value_local = dict_of_reg_value_CRD.copy()
        CR_d = {
            'General_information': {
                'CR_number': None, 'CR_coordinator': None,
                'CR_number_int': None, 'Organization': None, 'Initiator': None,
                'TDD_influece': None, 'Document_type': None, 'CR_reason': None,
                'Descr_tech_sol': None, 'NSC_category': None,
                'Impact_1_2_3': None,
                'Impact_DSA': None,
                'Impact_Evaluation': {
                    'Detailed_Design': 'NO',
                    'Licensing_Documentation': 'NO',
                    'Structural_Reliability': 'NO',
                    'Industrial_Safety': 'NO',
                    'Ecology_Impact': 'NO',
                    'Fire_Safety': 'NO',
                    'Schedule_Impact': 'NO',
                    'Cost_Impact': 'NO',
                    'Contract_Impact': 'NO'
                },
                'Method_CR': None, 'Comment': None, 'Contract': None,
                'Impact_cost': None, 'Schedule': None,
                'Comment_nont': None, 'Reg_date': None,
                'Constr_facility': None, 'Reason_code': None,
                'Change_type': None
            },
            'Confirmation': {}, 'Approval': {}, 'Supp_descr_docs': {},
            'TDD': {}, 'SSC': {},
            'List of documents proposed change': {}
        }

        parse_impact_evaluation(cells, CR_d)

        section = 'General'
        prev_section = 'General'
        _temp_dict = {}

        for row in range(first_cell[0], last_cell[0] + 1):
            if list(filter(
                    lambda x: x[0] == row and x[1] > last_cell[1] and cells[x],
                    cells)):
                continue

            if not any(
                    [cells.get((row, i)) for i in range(1, last_cell[1] + 1)]):
                continue

            first_text = None
            first_col = 1
            for col in range(1, last_cell[1] + 1):
                if cells.get((row, col)):
                    first_text = str(cells.get((row, col)))
                    first_col = col
                    break

            if first_text:
                header_text = first_text
                if re_var(['supporting', 'descr'], header_text) or re_var(
                        ['support', 'file'], header_text) or re_var(
                    ['melléklet'], header_text):
                    section = 'Sup_doc'
                elif re_var(['list', 'document', 'proposed'],
                            header_text) or re_var(
                    ['módosítását', 'javasolják'], header_text) or re_var(
                    ['dokumentumok', 'listája'], header_text):
                    section = 'Proposed_docs'
                elif re_var(['impact', 'init', 'TDD'], header_text):
                    section = 'TDD'
                elif re_var(['impact', 'TDD'], header_text):
                    if section == 'Sup_doc':
                        section = 'TDD'
                    elif section == 'TDD':
                        section = 'Other_TDD'
                elif re_var(['ssc'], header_text):
                    section = 'SSC'
                elif re_var(['final', 'eval'], header_text):
                    section = 'Final'
                elif re_var(['non-tech'], header_text) or re_var(['nontech'],
                                                                 header_text):
                    section = 'Nontech'
                elif re_var(['confirmat'], header_text):
                    section = 'Confirmation'
                elif re_var(['approv'], header_text) or re_var(['agreed'],
                                                               header_text) or re_var(
                    ['согласовано'], header_text) or re_var(['signat'],
                                                            header_text):
                    section = 'Approval'

            if section != prev_section:
                _temp_dict = {}
                prev_section = section

            is_bottom_up_sig = False
            c_date = None
            c_pos = None

            row_str = " ".join([str(cells.get((row, c))).lower() for c in
                                range(1, last_cell[1] + 1) if
                                cells.get((row, c))])

            if ('signature' in row_str or 'aláírás' in row_str) and (
                    'position' in row_str or 'beosztás' in row_str):
                for col in range(1, last_cell[1] + 1):
                    val = str(cells.get((row, col))).lower()
                    if 'dd.mm.yyyy' in val or 'dátuma' in val or 'date' in val:
                        c_date = col
                    elif 'position' in val or 'beosztás' in val or 'name' in val or 'név' in val:
                        c_pos = col
                    elif 'signature' in val or 'aláírás' in val:
                        is_bottom_up_sig = True

                if is_bottom_up_sig and c_pos:
                    role_text = cells.get((row - 1, 1))
                    if not role_text:
                        role_text = cells.get((row - 1, 2))

                    date_val = cells.get((row - 1, c_date)) if c_date else None
                    pos_name_val = cells.get((row - 1, c_pos))

                    if role_text and str(role_text).strip() not in ['None',
                                                                    '']:
                        role_str = str(role_text).replace('\n', ' ').strip()
                        if re_var(['init'], role_str):
                            if pos_name_val:
                                CR_d['General_information']['Initiator'] = str(
                                    pos_name_val)
                        else:
                            CR_d['Approval'][role_str] = {
                                'person': pos_name_val, 'date': date_val}
                    continue

            if section in ['TDD', 'Other_TDD', 'SSC', 'Confirmation',
                           'Approval', 'Proposed_docs']:
                if section == 'Approval' and first_text and (
                        re_var(['close'], first_text) or re_var(
                    ['end', 'form'], first_text)):
                    break

                if first_text and (
                        re_var(['code'], first_text) or re_var(['organ'],
                                                               first_text) or re_var(
                    ['posit'], first_text) or re_var(['role'],
                                                     first_text)):
                    _temp_dict = temp_dict_of_row(row, cells)
                    continue

                if not _temp_dict: continue

                code = None
                if section in ['TDD', 'Other_TDD', 'SSC', 'Proposed_docs']:
                    code_keys = list(
                        filter(lambda x: re_var(['code'],
                                                _temp_dict[x]) or re_var(
                            ['kód'], _temp_dict[x]),
                               _temp_dict))
                    if code_keys: code = cells.get((row, code_keys[0]))
                elif section == 'Confirmation':
                    code_keys = list(
                        filter(lambda x: re_var(['organ'], _temp_dict[x]),
                               _temp_dict))
                    if code_keys: code = cells.get((row, code_keys[0]))
                elif section == 'Approval':
                    code_keys = list(
                        filter(lambda x: re_var(['posit'],
                                                _temp_dict[x]) or re_var(
                            ['code'], _temp_dict[x]),
                               _temp_dict))
                    if code_keys: code = cells.get((row, code_keys[0]))

                if not code:
                    continue

                if section == 'SSC':
                    CR_d['SSC'][code] = {}
                elif section == 'Confirmation':
                    if code not in CR_d['Confirmation'].keys():
                        CR_d['Confirmation'][code] = {}
                elif section == 'Approval':
                    if code not in CR_d['Approval'].keys():
                        CR_d['Approval'][code] = {}
                elif section == 'Proposed_docs':
                    if code not in CR_d[
                        'List of documents proposed change'].keys():
                        CR_d['List of documents proposed change'][code] = {}
                else:
                    if code not in CR_d['TDD'].keys():
                        CR_d['TDD'][code] = {}

                for column_number in _temp_dict:
                    insert_value = cells.get((row, column_number))

                    if section == 'Proposed_docs' and insert_value:
                        if re_var(['revision'],
                                  _temp_dict[column_number]) or re_var(
                            ['revízió'],
                            _temp_dict[column_number]) or re_var(['rev'],
                                                                 _temp_dict[
                                                                     column_number]):
                            if '_' in str(insert_value):
                                CR_d['List of documents proposed change'][
                                    code]['Revision'] = \
                                    str(insert_value).split('_')[0]
                            else:
                                CR_d['List of documents proposed change'][
                                    code]['Revision'] = insert_value
                        elif re_var(['name'],
                                    _temp_dict[column_number]) or re_var(
                            ['megnevezés'],
                            _temp_dict[column_number]) or re_var(['cím'],
                                                                 _temp_dict[
                                                                     column_number]):
                            CR_d['List of documents proposed change'][code][
                                'Name'] = insert_value

                    elif section in ['TDD', 'Other_TDD'] and insert_value:
                        if re_var(['organ'], _temp_dict[column_number]):
                            CR_d['TDD'][code][
                                'Organization'] = insert_value
                        elif re_var(['new', 'revision'],
                                    _temp_dict[column_number]):
                            _temp_bool = True if 'es' in str(
                                insert_value).lower() else False
                            CR_d['TDD'][code]['New_rev_req'] = _temp_bool
                        elif re_var(['revision'],
                                    _temp_dict[column_number]):
                            if '_' in str(insert_value):
                                CR_d['TDD'][code]['Revision'] = \
                                    str(insert_value).split('_')[0]
                                CR_d['TDD'][code]['Version'] = \
                                    str(insert_value).split('_')[1]
                            else:
                                CR_d['TDD'][code][
                                    'Revision'] = insert_value
                                CR_d['TDD'][code]['Version'] = 0
                        elif re_var(['name'], _temp_dict[column_number]):
                            CR_d['TDD'][code]['Name'] = insert_value
                        elif re_var(['state'], _temp_dict[column_number]):
                            CR_d['TDD'][code]['Status'] = insert_value
                        elif re_var(['description'],
                                    _temp_dict[column_number]):
                            if 'Description' not in CR_d['TDD'][
                                code].keys():
                                CR_d['TDD'][code]['Description'] = [
                                    insert_value]
                            else:
                                CR_d['TDD'][code]['Description'].append(
                                    insert_value)
                        elif re_var(['impact'], _temp_dict[column_number]):
                            list_of_factors = ''
                            for factor in range(5):
                                _temp_bin = '0' if re_var(['no'],
                                                          cells.get((row,
                                                                     column_number + factor))) else '1'
                                list_of_factors += _temp_bin
                            CR_d['TDD'][code]['Impact'] = int(
                                list_of_factors, 2)
                    elif section == 'SSC':
                        if re_var(['name'], _temp_dict[column_number]):
                            CR_d['SSC'][code]['Name'] = insert_value
                        elif re_var(['description'],
                                    _temp_dict[column_number]):
                            CR_d['SSC'][code]['Description'] = insert_value
                    elif section == 'Confirmation':
                        if re_var(['posit'], _temp_dict[column_number]):
                            position = insert_value
                        elif re_var(['resp', 'pers'],
                                    _temp_dict[column_number]) or re_var(
                            ['name'], _temp_dict[column_number]):
                            name = insert_value
                            CR_d['Confirmation'][code][name] = {
                                'Position': None, 'Date': None}
                        elif re_var(['date'], _temp_dict[column_number]):
                            date = insert_value
                            if 'name' in locals() and name in \
                                    CR_d['Confirmation'][code]:
                                CR_d['Confirmation'][code][name][
                                    'Position'] = locals().get('position')
                                CR_d['Confirmation'][code][name]['Date'] = date
                    elif section == 'Approval':
                        if re_var(['person'],
                                  _temp_dict[column_number]) or re_var(
                            ['name'], _temp_dict[column_number]):
                            CR_d['Approval'][code]['person'] = insert_value
                        elif re_var(['date'], _temp_dict[column_number]):
                            CR_d['Approval'][code]['date'] = insert_value

            elif section == 'Sup_doc':
                if first_text and not (
                        re_var(['file', 'name'], first_text) or re_var(
                    ['details', 'document'], first_text) or re_var(
                    ['dokumentumok', 'adatai'], first_text)):
                    curr_cell, val = first_not_empty((row, first_col + 1),
                                                     cells, 'row')
                    if val:
                        CR_d['Supp_descr_docs'][first_text] = val

            elif section in ['General', 'Final', 'Nontech']:
                if section == 'General' and first_text and (
                        re_var(['descr', 'sol'], first_text) or re_var(
                    ['descr', 'change'], first_text) or re_var(
                    ['módosítások', 'leírása'], first_text)):
                    CR_d['General_information']['Descr_tech_sol'] = \
                        first_not_empty((row + 1, first_col), cells, 'row')[1]
                    continue

                _temp_dict2 = temp_dict_of_row(row, cells)
                if not _temp_dict2:
                    continue

                vals = list(_temp_dict2.values())

                if section == 'General':
                    for i, v in enumerate(vals):
                        v_str = str(v).lower()
                        if (
                                'number of building' in v_str or 'építmény száma' in v_str):
                            if i + 1 < len(vals) and vals[i + 1]:
                                bld_val = str(vals[i + 1]).strip()
                                if bld_val and bld_val not in CR_d['SSC']:
                                    CR_d['SSC'][bld_val] = {}

                if len(vals) >= 3 and (
                        re_var(['registr'], vals[0]) or re_var(['bejegyz'],
                                                               vals[
                                                                   0]) or re_var(
                    ['change', 'request'], vals[0])):
                    CR_d['General_information']['CR_number'] = vals[1]
                    CR_d['General_information']['Reg_date'] = vals[2]
                    continue

                sorted_keys = sorted(_temp_dict2.keys())
                paired_dict = {}

                idx = 0
                while idx < len(sorted_keys):
                    k1 = sorted_keys[idx]
                    v1 = _temp_dict2[k1]

                    is_h1 = any(
                        header_in_reg(v1, pos, dict_of_reg_value_local) for pos
                        in dict_of_reg_value_local)

                    if not is_h1:
                        idx += 1
                        continue

                    if idx + 1 < len(sorted_keys):
                        k2 = sorted_keys[idx + 1]
                        v2 = _temp_dict2[k2]
                        is_h2 = any(
                            header_in_reg(v2, pos, dict_of_reg_value_local) for
                            pos in dict_of_reg_value_local)

                        if is_h2:
                            paired_dict[v1] = None
                            idx += 1
                        else:
                            paired_dict[v1] = v2
                            idx += 2
                    else:
                        paired_dict[v1] = None
                        idx += 1

                for header, value in paired_dict.items():
                    for position in list(dict_of_reg_value_local.keys()):
                        if header_in_reg(header, position,
                                         dict_of_reg_value_local):
                            CR_d['General_information'][position] = value
                            dict_of_reg_value_local.pop(position)
                            break

    elif doc_mode == 'FCR':
        dict_of_reg_value_local = dict_of_reg_value_FCR.copy()
        CR_d = {
            'General_information': {
                'CR_reason': None, 'CR_number': None, 'Reg_date': None,
                'CR_coordinator': None, 'CR_number_int': None,
                'Organization': None,
                'Initiator': None, 'Initiator_pos': None,
                'Document_type': None,
                'Change_type': None, 'Activity_type': None,
                'Constr_facility': None,
                'Ini_Method_CR': None, 'Method_CR': None,
                'Ini_Method_justif': None,
                'Method_justif': None, 'Change_equipment': None,
                'Reason_code': None,
                'Descr_tech_sol': None, 'Final_status': None,
                'Impact_DSA': None,
                'Impact_Evaluation': {
                    'Detailed_Design': 'NO',
                    'Licensing_Documentation': 'NO',
                    'Structural_Reliability': 'NO',
                    'Industrial_Safety': 'NO',
                    'Ecology_Impact': 'NO',
                    'Fire_Safety': 'NO',
                    'Schedule_Impact': 'NO',
                    'Cost_Impact': 'NO',
                    'Contract_Impact': 'NO'
                },
                'Evaluation': {
                    'Material_eq': None, 'REPLACE?': None,
                    'Reject_comment': None, 'Refuse_comment': None,
                    'JD': {}, 'Impact_DDD': None, 'Impact_LDD': None,
                    'Impact_cost': None, 'Schedule': None,
                    'Prompt_req': None, 'NS': None, 'FS': None, 'IS': None,
                    'ES': None, 'SS': None
                }
            },
            'Approval': [],
            'Supp_descr_docs': {},
            'TDD_sets': {},
            'SSC': {}
        }

        parse_impact_evaluation(cells, CR_d)

        section = 'General'
        prev_section = 'General'
        subsection = False
        _descr_flag = 0
        _temp_dict = {}

        for row in range(first_cell[0], last_cell[0] + 1):
            if list(filter(
                    lambda x: x[0] == row and x[1] > last_cell[1] and cells[x],
                    cells)):
                warning += 1
                continue

            if not any(
                    [cells.get((row, i)) for i in range(1, last_cell[1] + 1)]):
                if section == 'Concurrence':
                    section = 'Approval'
                continue

            first_text = None
            first_col = 1
            for col in range(1, last_cell[1] + 1):
                if cells.get((row, col)):
                    first_text = str(cells.get((row, col)))
                    first_col = col
                    break

            if first_text:
                if section == 'General' and re_var(['affect', 'ssc'],
                                                   first_text):
                    section = 'SSC'
                elif section == 'SSC' and re_var(['list', 'rel', 'doc'],
                                                 first_text):
                    section = 'TDD'
                elif section == 'TDD' and re_var(['supp', 'descr'],
                                                 first_text):
                    section = 'Sup_doc'
                elif section == 'Sup_doc' and re_var(['evaluat', 'of'],
                                                     first_text):
                    section = 'Evaluation'
                elif section == 'Evaluation' and re_var(['linc', 'doc'],
                                                        first_text):
                    subsection = 'JD'
                elif section == 'Evaluation' and re_var(['concur', 'sheet'],
                                                        first_text):
                    section = 'Concurrence'
                elif section == 'Approval' and re_var(['close'], first_text):
                    section = 'Close'

                if section == 'SSC' and re_var(['not', 'spec'],
                                               first_text): continue
                if section == 'SSC' and re_var(['cod', 'SSC', 'n/a'],
                                               first_text): continue
                if section == 'TDD' and re_var(['set', 'which'],
                                               first_text): continue
                if section == 'Evaluation' and re_var(['type', 'of', 'change'],
                                                      first_text): continue
                if section == 'Evaluation' and re_var(['crit', 'imp'],
                                                      first_text): continue
                if section == 'Evaluation' and subsection == 'JD' and re_var(
                        ['file', 'extension'], first_text): continue
                if section == 'General' and re_var(['field', 'change', 'fcr'],
                                                   first_text): continue
                if section == 'General' and re_var(['change', 'init', 'for'],
                                                   first_text): continue

                if section == 'General' and _descr_flag == 1: _descr_flag = 0
                if section == 'Evaluation' and _descr_flag == 1: _descr_flag = 0

                if section == 'Approval' and re_var(['\\*', '\\*\\*'],
                                                    first_text, stop_words=[
                            '\\*\\*\\*']): continue
                if section == 'Sup_doc' and (
                        re_var(['file', 'ext'], first_text) or re_var(
                    ['end', 'init'], first_text)): continue

            if section != prev_section:
                _temp_dict = {}
                prev_section = section

            is_bottom_up_sig = False
            c_date = None
            c_pos = None

            row_str = " ".join([str(cells.get((row, c))).lower() for c in
                                range(1, last_cell[1] + 1) if
                                cells.get((row, c))])

            if ('signature' in row_str or 'aláírás' in row_str) and (
                    'position' in row_str or 'beosztás' in row_str):
                for col in range(1, last_cell[1] + 1):
                    val = str(cells.get((row, col))).lower()
                    if 'dd.mm.yyyy' in val or 'dátuma' in val or 'date' in val:
                        c_date = col
                    elif 'position' in val or 'beosztás' in val or 'name' in val or 'név' in val:
                        c_pos = col
                    elif 'signature' in val or 'aláírás' in val:
                        is_bottom_up_sig = True

                if is_bottom_up_sig and c_pos:
                    role_text = cells.get((row - 1, 1))
                    if not role_text:
                        role_text = cells.get((row - 1, 2))

                    date_val = cells.get((row - 1, c_date)) if c_date else None
                    pos_name_val = cells.get((row - 1, c_pos))

                    if role_text and str(role_text).strip() not in ['None',
                                                                    '']:
                        role_str = str(role_text).replace('\n', ' ').strip()
                        if re_var(['init'], role_str):
                            if pos_name_val:
                                CR_d['General_information']['Initiator'] = str(
                                    pos_name_val)
                        else:
                            CR_d['Approval'].append(
                                {'Position': role_str, 'Name': pos_name_val,
                                 'Date': date_val})
                    continue

            if section in ['TDD', 'SSC', 'Approval']:
                if section == 'Close' and first_text and re_var(
                        ['end', 'form'], first_text):
                    break

                if first_text and (
                        re_var(['code'], first_text) or re_var(['posit'],
                                                               first_text) or re_var(
                    ['role'], first_text)):
                    _temp_dict = temp_dict_of_row(row, cells)
                    continue

                if not _temp_dict:
                    continue

                code = None
                if section in ['TDD', 'SSC']:
                    code_keys = list(
                        filter(lambda x: re_var(['code'], _temp_dict[x]),
                               _temp_dict))
                    if code_keys: code = cells.get((row, code_keys[0]))
                elif section == 'Approval':
                    code_keys = list(
                        filter(lambda x: re_var(['posit'],
                                                _temp_dict[x]) or re_var(
                            ['code'], _temp_dict[x]),
                               _temp_dict))
                    if code_keys: code = cells.get((row, code_keys[0]))

                if not code:
                    continue

                if section == 'SSC':
                    CR_d['SSC'][code] = {}
                elif section == 'Approval':
                    CR_d['Approval'].append({'Position': code})
                else:
                    if code not in CR_d['TDD_sets'].keys():
                        CR_d['TDD_sets'][code] = {}

                for column_number in _temp_dict:
                    insert_value = cells.get((row, column_number))
                    if section == 'TDD' and insert_value:
                        if re_var(['set', 'revis'],
                                  _temp_dict[column_number]):
                            if '_' in str(insert_value):
                                CR_d['TDD_sets'][code]['Revision'] = \
                                    str(insert_value).split('_')[0]
                                CR_d['TDD_sets'][code]['Version'] = \
                                    str(insert_value).split('_')[1]
                            else:
                                CR_d['TDD_sets'][code][
                                    'Revision'] = insert_value
                                CR_d['TDD_sets'][code]['Version'] = 0
                        if re_var(['engin', 'cod'],
                                  _temp_dict[column_number]):
                            doc_code = insert_value
                            if 'Documents' not in CR_d['TDD_sets'][
                                code].keys():
                                CR_d['TDD_sets'][code]['Documents'] = [
                                    {'Code': doc_code}]
                            if doc_code not in list(
                                    map(lambda x: x['Code'],
                                        CR_d['TDD_sets'][code][
                                            'Documents'])):
                                CR_d['TDD_sets'][code]['Documents'].append(
                                    {'Code': doc_code})
                        if re_var(['engin', 'name'],
                                  _temp_dict[column_number]):
                            index_of_doc_code = [index for index, value in
                                                 enumerate(
                                                     CR_d['TDD_sets'][
                                                         code][
                                                         'Documents']) if
                                                 value[
                                                     'Code'] == doc_code][
                                0]
                            CR_d['TDD_sets'][code]['Documents'][
                                index_of_doc_code]['Name'] = insert_value
                        if re_var(['revis', 'ED'],
                                  _temp_dict[column_number]) or re_var(
                            ['ED', 'revis'],
                            _temp_dict[column_number]):
                            if '_' in str(insert_value):
                                CR_d['TDD_sets'][code]['Documents'][
                                    index_of_doc_code]['Revision'] = \
                                    str(insert_value).split('_')[0]
                                CR_d['TDD_sets'][code]['Documents'][
                                    index_of_doc_code]['Version'] = \
                                    str(insert_value).split('_')[1]
                            else:
                                CR_d['TDD_sets'][code]['Documents'][
                                    index_of_doc_code][
                                    'Revision'] = insert_value
                                CR_d['TDD_sets'][code]['Documents'][
                                    index_of_doc_code]['Version'] = 0
                        if re_var(['sheets'], _temp_dict[column_number]):
                            sheets = str(insert_value)
                            if 'Sheets' not in \
                                    CR_d['TDD_sets'][code]['Documents'][
                                        index_of_doc_code]:
                                CR_d['TDD_sets'][code]['Documents'][
                                    index_of_doc_code]['Sheets'] = {}
                            if sheets not in \
                                    CR_d['TDD_sets'][code]['Documents'][
                                        index_of_doc_code][
                                        'Sheets'].keys():
                                CR_d['TDD_sets'][code]['Documents'][
                                    index_of_doc_code]['Sheets'][
                                    sheets] = {}
                        if re_var(['chang', 'amx'],
                                  _temp_dict[column_number]):
                            if 'AMX' not in \
                                    CR_d['TDD_sets'][code]['Documents'][
                                        index_of_doc_code]['Sheets'][
                                        sheets].keys():
                                CR_d['TDD_sets'][code]['Documents'][
                                    index_of_doc_code]['Sheets'][sheets][
                                    'AMX'] = insert_value
                            else:
                                if insert_value not in \
                                        CR_d['TDD_sets'][code][
                                            'Documents'][
                                            index_of_doc_code]['Sheets'][
                                            sheets]['AMX']:
                                    CR_d['TDD_sets'][code]['Documents'][
                                        index_of_doc_code]['Sheets'][
                                        sheets][
                                        'AMX'] += f'\n{insert_value}'
                        if re_var(['description'],
                                  _temp_dict[column_number]):
                            if 'Description' not in \
                                    CR_d['TDD_sets'][code]['Documents'][
                                        index_of_doc_code]['Sheets'][
                                        sheets].keys():
                                CR_d['TDD_sets'][code]['Documents'][
                                    index_of_doc_code]['Sheets'][sheets][
                                    'Description'] = [insert_value]
                            else:
                                CR_d['TDD_sets'][code]['Documents'][
                                    index_of_doc_code]['Sheets'][sheets][
                                    'Description'].append(insert_value)
                    if section == 'SSC':
                        if re_var(['system', 'code'],
                                  _temp_dict[column_number]):
                            CR_d['SSC'][code]['Sys_code'] = insert_value
                        elif re_var(['component'],
                                    _temp_dict[column_number]):
                            CR_d['SSC'][code]['Component'] = insert_value
                    if section == 'Approval':
                        if re_var(['person'],
                                  _temp_dict[column_number]) or re_var(
                            ['name'], _temp_dict[column_number]):
                            CR_d['Approval'][-1]['Name'] = insert_value
                        elif re_var(['date'], _temp_dict[column_number]):
                            CR_d['Approval'][-1]['Date'] = insert_value

            elif section == 'Sup_doc' or (
                    doc_mode == 'FCR' and section == 'Evaluation' and subsection == 'JD'):
                if first_text and not (
                        re_var(['file', 'name'], first_text) or re_var(
                    ['details', 'document'], first_text) or re_var(
                    ['dokumentumok', 'adatai'], first_text)):
                    if section == 'Sup_doc':
                        if first_text in CR_d['Supp_descr_docs'].keys():
                            curr_cell, val = first_not_empty(
                                (row, first_col + 1), cells, 'row')
                            if val:
                                CR_d['Supp_descr_docs'][first_text][
                                    'Title'].append(val)
                        else:
                            curr_cell, val = first_not_empty(
                                (row, first_col + 1), cells, 'row')
                            if val:
                                CR_d['Supp_descr_docs'][first_text] = {
                                    'Title': [val]}
                    if section == 'Evaluation' and subsection == 'JD':
                        if 'JD' not in CR_d['General_information'][
                            'Evaluation']:
                            CR_d['General_information']['Evaluation'][
                                'JD'] = {}
                        curr_cell, \
                            CR_d['General_information']['Evaluation']['JD'][
                                first_text] = first_not_empty(
                            (row, first_col + 1),
                            cells, 'row')
                        if re_var(['type', 'change'], first_text):
                            subsection = False

            elif section in ['General', 'Concurrence', 'Close'] or (
                    section == 'Evaluation' and subsection != 'JD'):
                if section == 'General' and first_text and re_var(
                        ['descrip', 'engin', 'change'], first_text):
                    CR_d['General_information']['Descr_tech_sol'] = \
                        first_not_empty((row + 1, 1), cells, 'row')[1]
                    _descr_flag = 1
                if section == 'Evaluation' and first_text and re_var(
                        ['comment', 'reason', 'reject'], first_text):
                    CR_d['General_information']['Evaluation'][
                        'Reject_comment'] = \
                        first_not_empty((row + 1, 1), cells, 'row')[1]
                    _descr_flag = 1
                if section == 'Evaluation' and first_text and re_var(
                        ['comment', 'reason', 'refus'], first_text):
                    CR_d['General_information']['Evaluation'][
                        'Refuse_comment'] = \
                        first_not_empty((row + 1, 1), cells, 'row')[1]
                    _descr_flag = 1

                _temp_dict2 = temp_dict_of_row(row, cells)
                if not _temp_dict2:
                    continue

                sorted_keys = sorted(_temp_dict2.keys())
                paired_dict = {}

                idx = 0
                while idx < len(sorted_keys):
                    k1 = sorted_keys[idx]
                    v1 = _temp_dict2[k1]

                    is_h1 = any(
                        header_in_reg(v1, pos, dict_of_reg_value_local) for pos
                        in dict_of_reg_value_local)

                    if not is_h1:
                        idx += 1
                        continue

                    if idx + 1 < len(sorted_keys):
                        k2 = sorted_keys[idx + 1]
                        v2 = _temp_dict2[k2]
                        is_h2 = any(
                            header_in_reg(v2, pos, dict_of_reg_value_local) for
                            pos in dict_of_reg_value_local)

                        if is_h2:
                            paired_dict[v1] = None
                            idx += 1
                        else:
                            paired_dict[v1] = v2
                            idx += 2
                    else:
                        paired_dict[v1] = None
                        idx += 1

                for header, value in paired_dict.items():
                    for position in list(dict_of_reg_value_local.keys()):
                        if header_in_reg(header, position,
                                         dict_of_reg_value_local):
                            if doc_mode == 'FCR' and section == 'Evaluation':
                                CR_d['General_information']['Evaluation'][
                                    position] = value
                            else:
                                CR_d['General_information'][position] = value
                            dict_of_reg_value_local.pop(position)
                            break

    elif doc_mode == 'CR':
        dict_of_reg_value_local = dict_of_reg_value_CR.copy()

        CR_d = {
            'General_information': {
                'CR_number': None, 'Reg_date': None,
                'Initiator': None, 'Organization': None, 'Responsible': None,
                'Impact_DSA': None,
                'Impact_Evaluation': {
                    'Detailed_Design': 'NO',
                    'Licensing_Documentation': 'NO',
                    'Structural_Reliability': 'NO',
                    'Industrial_Safety': 'NO',
                    'Ecology_Impact': 'NO',
                    'Fire_Safety': 'NO',
                    'Schedule_Impact': 'NO',
                    'Cost_Impact': 'NO',
                    'Contract_Impact': 'NO'
                },
                'CR_coordinator': None, 'CR_number_int': None,
                'Contract': None,
                'Impact_cost': None, 'Schedule': None,
                'Comment_nont': None, 'Constr_facility': None,
                'Reason_code': None, 'CR_reason': None, 'Descr_tech_sol': None,
                'Document_type': None, 'Change_type': None
            },
            'Confirmation': {}, 'Approval': {}, 'Configur': {}, 'TDD': {},
            'SSC': {},
            'List of documents proposed change': {}, 'Supp_descr_docs': {}
        }

        parse_impact_evaluation(cells, CR_d)

        section = 'General'
        prev_section = 'General'
        _temp_dict = {}

        for row in range(first_cell[0], last_cell[0] + 1):
            if list(filter(
                    lambda x: x[0] == row and x[1] > last_cell[1] and cells[x],
                    cells)):
                continue

            if not any(
                    [cells.get((row, i)) for i in range(1, last_cell[1] + 1)]):
                continue

            first_text = None
            first_col = 1
            for col in range(1, last_cell[1] + 1):
                if cells.get((row, col)):
                    first_text = str(cells.get((row, col)))
                    first_col = col
                    break

            if first_text:
                header_text = first_text
                if re_var(['supporting', 'descr'], header_text) or re_var(
                        ['support', 'file'], header_text) or re_var(
                    ['melléklet'], header_text) or re_var(['csatolva'],
                                                          header_text):
                    section = 'Sup_doc'
                elif re_var(['init', 'item'], header_text) or re_var(
                        ['scope', 'change'], header_text):
                    section = 'TDD'
                elif re_var(['list', 'document', 'proposed'],
                            header_text) or re_var(
                    ['módosítását', 'javasolják'], header_text) or re_var(
                    ['dokumentumok', 'listája'], header_text):
                    section = 'Proposed_docs'
                elif re_var(['conf', 'item'], header_text):
                    section = 'Configur'
                elif re_var(['affect', 'syst'], header_text) or re_var(
                        ['affect', 'ssc'], header_text) or re_var(
                    ['changed', 'ssc'], header_text):
                    section = 'SSC'
                elif re_var(['confirmat'], header_text):
                    section = 'Confirmation'
                elif re_var(['non-tech'], header_text) or re_var(['nontech'],
                                                                 header_text):
                    section = 'Nontech'
                elif re_var(['approv'], header_text) or re_var(['agreed'],
                                                               header_text) or re_var(
                    ['согласовано'], header_text) or re_var(['signat'],
                                                            header_text):
                    section = 'Approval'

            if section != prev_section:
                _temp_dict = {}
                prev_section = section

            is_bottom_up_sig = False
            c_date = None
            c_pos = None

            row_str = " ".join([str(cells.get((row, c))).lower() for c in
                                range(1, last_cell[1] + 1) if
                                cells.get((row, c))])

            if ('signature' in row_str or 'aláírás' in row_str) and (
                    'position' in row_str or 'beosztás' in row_str):
                for col in range(1, last_cell[1] + 1):
                    val = str(cells.get((row, col))).lower()
                    if 'dd.mm.yyyy' in val or 'dátuma' in val or 'date' in val:
                        c_date = col
                    elif 'position' in val or 'beosztás' in val or 'name' in val or 'név' in val:
                        c_pos = col
                    elif 'signature' in val or 'aláírás' in val:
                        is_bottom_up_sig = True

                if is_bottom_up_sig and c_pos:
                    role_text = cells.get((row - 1, 1))
                    if not role_text:
                        role_text = cells.get((row - 1, 2))

                    date_val = cells.get((row - 1, c_date)) if c_date else None
                    pos_name_val = cells.get((row - 1, c_pos))

                    if role_text and str(role_text).strip() not in ['None',
                                                                    '']:
                        role_str = str(role_text).replace('\n', ' ').strip()
                        if re_var(['init'], role_str):
                            if pos_name_val:
                                CR_d['General_information']['Initiator'] = str(
                                    pos_name_val)
                        else:
                            CR_d['Approval'][role_str] = {
                                'person': pos_name_val, 'date': date_val}
                    continue

            if section in ['TDD', 'Configur', 'SSC', 'Confirmation',
                           'Approval', 'Proposed_docs']:
                if section == 'Approval' and first_text and (
                        re_var(['close'], first_text) or re_var(
                    ['end', 'form'], first_text)):
                    break

                if first_text and (
                        re_var(['code'], first_text) or re_var(['posit'],
                                                               first_text) or re_var(
                    ['role'], first_text)):
                    _temp_dict = temp_dict_of_row(row, cells)
                    if section == 'SSC':
                        not_empty = \
                            first_not_empty((row, first_col + 1), cells,
                                            'row')[0]
                        if not_empty:
                            cell_of_DSA = \
                                first_not_empty(
                                    (not_empty[0], not_empty[1] + 1),
                                    cells, 'row')[1]
                            CR_d['General_information'][
                                'Impact_DSA'] = cell_of_DSA
                    continue

                if not _temp_dict: continue

                code = None
                if section in ['TDD', 'Configur', 'SSC', 'Proposed_docs']:
                    code_keys = list(
                        filter(lambda x: re_var(['code'],
                                                _temp_dict[x]) or re_var(
                            ['kód'], _temp_dict[x]),
                               _temp_dict))
                    if code_keys: code = cells.get((row, code_keys[0]))

                elif section in ['Approval', 'Confirmation']:
                    code_keys = list(
                        filter(lambda x: re_var(['posit'],
                                                _temp_dict[x]) or re_var(
                            ['code'], _temp_dict[x]),
                               _temp_dict))
                    if code_keys: code = cells.get((row, code_keys[0]))

                if section in ['Approval', 'Confirmation']:
                    for column_number in _temp_dict:
                        if re_var(['init'], _temp_dict[column_number]):
                            insert_value = cells.get((row, column_number))
                            if insert_value:
                                val = str(insert_value)
                                pos_val = cells.get(
                                    (row, code_keys[0])) if code_keys else ""
                                if pos_val and str(pos_val).strip():
                                    val = f"{pos_val} - {val}"
                                if not CR_d['General_information'][
                                    'Initiator']:
                                    CR_d['General_information'][
                                        'Initiator'] = val
                                elif val not in CR_d['General_information'][
                                    'Initiator']:
                                    CR_d['General_information'][
                                        'Initiator'] += f"\n{val}"

                if not code:
                    continue

                if section == 'SSC':
                    CR_d['SSC'][code] = {}
                elif section == 'Confirmation':
                    if code not in CR_d['Confirmation'].keys():
                        CR_d['Confirmation'][code] = {}
                elif section == 'Approval':
                    if code not in CR_d['Approval'].keys():
                        CR_d['Approval'][code] = {}
                elif section == 'Proposed_docs':
                    if code not in list(
                            CR_d['List of documents proposed change'].keys()):
                        CR_d['List of documents proposed change'][code] = {}
                elif section == 'Configur':
                    CR_d['Configur'][code] = {}
                else:
                    if code not in list(CR_d['TDD'].keys()):
                        CR_d['TDD'][code] = {}

                for column_number in _temp_dict:
                    insert_value = cells.get((row, column_number))

                    if section == 'Proposed_docs' and insert_value:
                        if re_var(['revision'],
                                  _temp_dict[column_number]) or re_var(
                            ['revízió'],
                            _temp_dict[column_number]) or re_var(['rev'],
                                                                 _temp_dict[
                                                                     column_number]):
                            if '_' in str(insert_value):
                                CR_d['List of documents proposed change'][
                                    code]['Revision'] = \
                                    str(insert_value).split('_')[0]
                            else:
                                CR_d['List of documents proposed change'][
                                    code]['Revision'] = insert_value
                        elif re_var(['name'],
                                    _temp_dict[column_number]) or re_var(
                            ['megnevezés'],
                            _temp_dict[column_number]) or re_var(['cím'],
                                                                 _temp_dict[
                                                                     column_number]):
                            CR_d['List of documents proposed change'][code][
                                'Name'] = insert_value

                    elif section in ['TDD', 'Configur'] and insert_value:
                        if re_var(['revision'], _temp_dict[column_number]):
                            if '_' in str(insert_value):
                                CR_d[section][code]['Revision'] = \
                                    str(insert_value).split('_')[0]
                                CR_d[section][code]['Version'] = \
                                    str(insert_value).split('_')[1]
                            else:
                                CR_d[section][code]['Revision'] = insert_value
                                CR_d[section][code]['Version'] = 0
                        elif re_var(['name'], _temp_dict[column_number]):
                            CR_d[section][code]['Name'] = insert_value
                        elif re_var(['state'], _temp_dict[column_number]):
                            CR_d[section][code]['Status'] = insert_value
                        elif re_var(['description'],
                                    _temp_dict[column_number]):
                            if 'Description' not in list(
                                    CR_d[section][code].keys()):
                                CR_d[section][code][
                                    'Description'] = insert_value
                            else:
                                CR_d[section][code][
                                    'Description'] += insert_value
                        elif re_var(['reason'], _temp_dict[column_number]):
                            if 'Merged' in str(
                                    type(ws._cells[row, column_number])):
                                merged_cells = list(filter(lambda
                                                               x: x.min_col <= column_number <= x.max_col and x.min_row <= row <= x.max_row,
                                                           list(
                                                               ws.merged_cells.ranges)))[
                                    0]
                                insert_value = cells[
                                    merged_cells.min_row, merged_cells.min_col]
                            CR_d[section][code]['Reason'] = insert_value
                        elif re_var(['eval, imp'],
                                    _temp_dict[column_number]):
                            CR_d[section][code]['Imp_eval'] = insert_value
                    elif section == 'SSC':
                        if re_var(['description'], _temp_dict[column_number]):
                            CR_d['SSC'][code]['Description'] = insert_value

                    elif section == 'Confirmation':
                        if re_var(['posit'], _temp_dict[column_number]):
                            position = insert_value
                        elif re_var(['resp', 'pers'],
                                    _temp_dict[column_number]) or re_var(
                            ['name'], _temp_dict[column_number]) or re_var(
                            ['approv'], _temp_dict[column_number]):
                            name = insert_value
                            CR_d['Confirmation'][code][name] = {
                                'Position': None, 'Date': None}
                        elif re_var(['date'], _temp_dict[column_number]):
                            date = insert_value
                            if 'name' in locals() and name in \
                                    CR_d['Confirmation'][code]:
                                CR_d['Confirmation'][code][name][
                                    'Position'] = locals().get('position')
                                CR_d['Confirmation'][code][name]['Date'] = date

                    elif section == 'Approval':
                        if re_var(['person'],
                                  _temp_dict[column_number]) or re_var(
                            ['name'], _temp_dict[column_number]) or re_var(
                            ['approv'], _temp_dict[column_number]):
                            CR_d['Approval'][code]['person'] = insert_value
                        elif re_var(['date'], _temp_dict[column_number]):
                            CR_d['Approval'][code]['date'] = insert_value

            elif section == 'Sup_doc':
                if first_text and not (
                        re_var(['file', 'name'], first_text) or re_var(
                    ['details', 'document'], first_text) or re_var(
                    ['dokumentumok', 'adatai'], first_text)):
                    curr_cell, val = first_not_empty((row, first_col + 1),
                                                     cells, 'row')
                    if val:
                        CR_d['Supp_descr_docs'][first_text] = val

            elif section in ['General', 'Nontech']:
                if section == 'General' and first_text and (
                        re_var(['descr', 'sol'], first_text) or re_var(
                    ['descr', 'change'], first_text) or re_var(
                    ['módosítások', 'leírása'], first_text)):
                    CR_d['General_information']['Descr_tech_sol'] = \
                        first_not_empty((row + 1, first_col), cells, 'row')[1]
                    continue

                _temp_dict2 = temp_dict_of_row(row, cells)
                if not _temp_dict2:
                    continue

                vals = list(_temp_dict2.values())

                if section == 'General':
                    for i, v in enumerate(vals):
                        v_str = str(v).lower()
                        if (
                                'number of building' in v_str or 'építmény száma' in v_str):
                            if i + 1 < len(vals) and vals[i + 1]:
                                bld_val = str(vals[i + 1]).strip()
                                if bld_val and bld_val not in CR_d['SSC']:
                                    CR_d['SSC'][bld_val] = {}

                if len(vals) == 3 and (
                        re_var(['registr'], vals[0]) or re_var(['bejegyz'],
                                                               vals[
                                                                   0]) or re_var(
                    ['change', 'request'], vals[0])):
                    CR_d['General_information']['CR_number'] = vals[1]
                    CR_d['General_information']['Reg_date'] = vals[2]
                    continue

                sorted_keys = sorted(_temp_dict2.keys())
                paired_dict = {}

                idx = 0
                while idx < len(sorted_keys):
                    k1 = sorted_keys[idx]
                    v1 = _temp_dict2[k1]

                    is_h1 = any(
                        header_in_reg(v1, pos, dict_of_reg_value_local) for pos
                        in dict_of_reg_value_local)

                    if not is_h1:
                        idx += 1
                        continue

                    if idx + 1 < len(sorted_keys):
                        k2 = sorted_keys[idx + 1]
                        v2 = _temp_dict2[k2]
                        is_h2 = any(
                            header_in_reg(v2, pos, dict_of_reg_value_local) for
                            pos in dict_of_reg_value_local)

                        if is_h2:
                            paired_dict[v1] = None
                            idx += 1
                        else:
                            paired_dict[v1] = v2
                            idx += 2
                    else:
                        paired_dict[v1] = None
                        idx += 1

                for header, value in paired_dict.items():
                    for position in list(dict_of_reg_value_local.keys()):
                        if header_in_reg(header, position,
                                         dict_of_reg_value_local):
                            CR_d['General_information'][position] = value
                            dict_of_reg_value_local.pop(position)
                            break

    print(f'>>> Файл {short_table_name} успешно разобран. Формируется JSON...')
    return CR_d


NS_template = {
    'File_name': str, 'Change_request_No': str, 'Reg_date': str,
    'Contr_change_coord': str,
    'Type_of_changes': str, 'Constr_facility': str, 'Type_of_changed_doc': str,
    'E-log?': bool,
    'Ini_org': str, 'Change_ini': str, 'Ini_internal_CR': str,
    'Cod_of_reason': str,
    'Other_reason': bool, 'Descr_of_change': str, 'Rel_to_prev_CR': str,
    'Approval_method': str,
    'Just_simple': str, 'Signature_list': [], 'TDD': [], 'SSC': [],
    'Support_files': []
}


def dicts_normalization(original_dict_name):
    normalized_dict = copy.deepcopy(NS_template)
    original_dict = result[original_dict_name].copy()
    normalized_dict['File_name'] = original_dict_name

    if 'TDD' in original_dict:
        normalized_dict['Change_request_No'] = original_dict[
            'General_information'].get('CR_number')
        normalized_dict['Reg_date'] = original_dict['General_information'].get(
            'Reg_date')
        normalized_dict['Contr_change_coord'] = original_dict[
            'General_information'].get('CR_coordinator')
        normalized_dict['Type_of_changes'] = original_dict[
            'General_information'].get('Change_type')
        normalized_dict['Constr_facility'] = original_dict[
            'General_information'].get('Constr_facility')
        normalized_dict['Type_of_changed_doc'] = original_dict[
            'General_information'].get('Document_type')
        normalized_dict['E-log?'] = None

        normalized_dict['Ini_org'] = original_dict['General_information'].get(
            'Organization')
        normalized_dict['Change_ini'] = original_dict[
            'General_information'].get('Initiator')
        normalized_dict['Ini_internal_CR'] = original_dict[
            'General_information'].get('CR_number_int')
        normalized_dict['Cod_of_reason'] = original_dict[
            'General_information'].get('Reason_code')

        other_reason = original_dict['General_information'].get('CR_reason')
        if not other_reason:
            other_reason = '\n'.join(
                filter(lambda x: x is not None, list(
                    map(lambda x: original_dict['TDD'][x].get(
                        'Reason') if 'Reason' in original_dict['TDD'][
                        x] else None,
                        original_dict['TDD']))))
        normalized_dict['Other_reason'] = other_reason

        normalized_dict['Descr_of_change'] = original_dict[
            'General_information'].get('Descr_tech_sol')

        normalized_dict['Rel_to_prev_CR'] = normalized_dict[
            'Approval_method'] = \
            normalized_dict['Just_simple'] = normalized_dict[
            'NSC_category'] = None
        normalized_dict['Impact_direct_123'] = True if original_dict[
            'SSC'] else False
        normalized_dict['Impact_DSA'] = original_dict[
            'General_information'].get(
            'Impact_DSA')
        normalized_dict['Impact_structural_geom'] = normalized_dict[
            'Impact_nucl'] = normalized_dict['Impact_fire'] = normalized_dict[
            'Impact_industrial'] = normalized_dict['Impact_environment'] = \
            normalized_dict['Impact_TDD'] = normalized_dict['Impact_lic_doc'] = \
            normalized_dict['Prompt_TDD?'] = normalized_dict[
            'Material_equivalent?'] = normalized_dict[
            'Comments_eng_eval'] = None
        normalized_dict['Impact_contract'] = \
            original_dict['General_information'].get('Contract')
        normalized_dict['Impact_cost'] = original_dict[
            'General_information'].get('Impact_cost')
        normalized_dict['Impact_schedule'] = original_dict[
            'General_information'].get('Schedule')
        normalized_dict['Comments_nont_ass'] = original_dict[
            'General_information'].get('Comment_nont')

        for key in original_dict['Confirmation']:
            _temp_dict = {}
            if original_dict['Confirmation'][key]:
                _temp_name = list(original_dict['Confirmation'][key].keys())[0]
                _temp_dict['Role'] = None
                _temp_dict['Position'] = key
                _temp_dict['Name_sur'] = _temp_name
                _temp_dict['FMV_number'] = _temp_dict['HAEA_reg'] = None
                _temp_dict['Date'] = \
                    original_dict['Confirmation'][key][_temp_name]['Date']
                normalized_dict['Signature_list'].append(_temp_dict)

        for key in original_dict['Approval']:
            _temp_dict = {}
            _temp_dict['Role'] = None
            _temp_dict['Position'] = key
            if isinstance(original_dict['Approval'][key], dict):
                _temp_dict['Name_sur'] = original_dict['Approval'][key].get(
                    'person')
                _temp_dict['Date'] = original_dict['Approval'][key].get('date')
            else:
                _temp_dict['Name_sur'] = original_dict['Approval'][key]
                _temp_dict['Date'] = None
            _temp_dict['FMV_number'] = _temp_dict['HAEA_reg'] = None
            normalized_dict['Signature_list'].append(_temp_dict)

        for key in original_dict['TDD']:
            _temp_dict = {}
            _temp_dict['Type'] = 'TDD'
            _temp_dict['Set_code'] = _temp_dict['Set_name'] = _temp_dict[
                'Set_rev'] = _temp_dict['Ser_version'] = _temp_dict[
                'Set_status'] = None
            _temp_dict['ED_code'] = key
            _temp_dict['ED_name'] = original_dict['TDD'][key].get('Name')
            _temp_dict['ED_rev'] = original_dict['TDD'][key].get('Revision')
            _temp_dict['ED_version'] = original_dict['TDD'][key].get('Version')
            _temp_dict['ED_status'] = original_dict['TDD'][key].get('Status')
            _temp_dict['Changed_sheets'] = _temp_dict['AMX_AM'] = None
            _temp_dict['Descr_of_change'] = original_dict['TDD'][key].get(
                'Description')
            _temp_dict['New_rev?'] = None
            normalized_dict['TDD'].append(_temp_dict)

        if 'Configur' in original_dict:
            for key in original_dict['Configur']:
                _temp_dict = {}
                _temp_dict['Type'] = 'Configur'
                _temp_dict['Set_code'] = _temp_dict['Set_name'] = _temp_dict[
                    'Set_rev'] = _temp_dict['Ser_version'] = _temp_dict[
                    'Set_status'] = None
                _temp_dict['ED_code'] = key
                _temp_dict['ED_name'] = original_dict['Configur'][key].get(
                    'Name')
                _temp_dict['ED_rev'] = original_dict['Configur'][key].get(
                    'Revision')
                _temp_dict['ED_version'] = original_dict['Configur'][key].get(
                    'Version')
                _temp_dict['ED_status'] = original_dict['Configur'][key].get(
                    'Status')
                _temp_dict['Changed_sheets'] = _temp_dict['AMX_AM'] = None
                _temp_dict['Descr_of_change'] = original_dict['Configur'][
                    key].get(
                    'Description')
                _temp_dict['New_rev?'] = original_dict['Configur'][key].get(
                    'Imp_eval')
                normalized_dict['TDD'].append(_temp_dict)

        for key in original_dict['SSC']:
            _temp_dict = {}
            _temp_dict['Type'] = 'Changed'
            _temp_dict['List_SSC'] = key
            _temp_dict['Name_KKS'] = None
            _temp_dict['Descr_of_change_KKS'] = original_dict['SSC'][key].get(
                'Description')
            normalized_dict['SSC'].append(_temp_dict)

        if 'Supp_descr_docs' in original_dict:
            for key in original_dict['Supp_descr_docs']:
                _temp_dict = {}
                _temp_dict['Type'] = 'Supporting'
                _temp_dict['File_name'] = key
                _temp_dict['File_content'] = original_dict['Supp_descr_docs'][
                    key]
                normalized_dict['Support_files'].append(_temp_dict)

    elif 'TDD_sets' in original_dict:
        normalized_dict['Change_request_No'] = \
            original_dict['General_information']['CR_number']
        normalized_dict['Reg_date'] = original_dict['General_information'][
            'Reg_date']
        normalized_dict['Contr_change_coord'] = \
            original_dict['General_information']['CR_coordinator']
        normalized_dict['Type_of_changes'] = \
            original_dict['General_information']['Change_type']
        normalized_dict['Constr_facility'] = \
            original_dict['General_information']['Constr_facility']
        normalized_dict['Type_of_changed_doc'] = \
            original_dict['General_information']['Document_type']
        normalized_dict['E-log?'] = None
        normalized_dict['Ini_org'] = original_dict['General_information'][
            'Organization']
        normalized_dict['Change_ini'] = original_dict['General_information'][
            'Initiator']
        normalized_dict['Ini_internal_CR'] = \
            original_dict['General_information']['CR_number_int']
        normalized_dict['Cod_of_reason'] = \
            original_dict['General_information']['Reason_code']
        normalized_dict['Other_reason'] = original_dict['General_information'][
            'CR_reason']
        normalized_dict['Descr_of_change'] = \
            original_dict['General_information']['Descr_tech_sol']
        normalized_dict['Rel_to_prev_CR'] = None
        normalized_dict['Approval_method'] = \
            original_dict['General_information']['Method_CR']
        normalized_dict['Just_simple'] = original_dict['General_information'][
            'Ini_Method_justif']
        normalized_dict['NSC_category'] = None
        normalized_dict['Impact_direct_123'] = True if original_dict[
            'SSC'] else False
        normalized_dict['Impact_DSA'] = None
        normalized_dict['Impact_structural_geom'] = \
            original_dict['General_information']['Evaluation']['SS']
        normalized_dict['Impact_nucl'] = \
            original_dict['General_information']['Evaluation']['NS']
        normalized_dict['Impact_fire'] = \
            original_dict['General_information']['Evaluation']['FS']
        normalized_dict['Impact_industrial'] = \
            original_dict['General_information']['Evaluation']['IS']
        normalized_dict['Impact_environment'] = \
            original_dict['General_information']['Evaluation']['ES']
        normalized_dict['Impact_TDD'] = \
            original_dict['General_information']['Evaluation']['Impact_DDD']
        normalized_dict['Impact_lic_doc'] = \
            original_dict['General_information']['Evaluation']['Impact_LDD']
        normalized_dict['Prompt_TDD?'] = \
            original_dict['General_information']['Evaluation']['Prompt_req']
        normalized_dict['Material_equivalent?'] = \
            original_dict['General_information']['Evaluation']['Material_eq']
        normalized_dict['Comments_eng_eval'] = smart_join([original_dict[
                                                               'General_information'][
                                                               'Evaluation'][
                                                               'Reject_comment'],
                                                           original_dict[
                                                               'General_information'][
                                                               'Evaluation'][
                                                               'Refuse_comment']])
        normalized_dict['Impact_contract'] = None
        normalized_dict['Impact_cost'] = \
            original_dict['General_information']['Evaluation']['Impact_cost']
        normalized_dict['Impact_schedule'] = \
            original_dict['General_information']['Evaluation']['Schedule']
        normalized_dict['Comments_nont_ass'] = None

        for key in original_dict['Approval']:
            _temp_dict = {}
            _temp_dict['Role'] = None
            _temp_dict['Position'] = key['Position']
            _temp_dict['Name_sur'] = key['Name']
            _temp_dict['FMV_number'] = _temp_dict['HAEA_reg'] = None
            _temp_dict['Date'] = key['Date']
            normalized_dict['Signature_list'].append(_temp_dict)

        for key in original_dict['TDD_sets']:
            for document in original_dict['TDD_sets'][key]['Documents']:
                _temp_dict = {}
                _temp_dict['Type'] = 'TDD'
                _temp_dict['Set_code'] = key
                _temp_dict['Set_name'] = None
                _temp_dict['Set_rev'] = original_dict['TDD_sets'][key][
                    'Revision']
                _temp_dict['Ser_version'] = original_dict['TDD_sets'][key][
                    'Version']
                _temp_dict['Set_status'] = None
                _temp_dict['ED_code'] = document
                _temp_dict['ED_name'] = \
                    original_dict['TDD_sets'][key]['Documents'][document][
                        'Name']
                _temp_dict['ED_rev'] = \
                    original_dict['TDD_sets'][key]['Documents'][document][
                        'Revision']
                _temp_dict['ED_version'] = \
                    original_dict['TDD_sets'][key]['Documents'][document][
                        'Version']
                _temp_dict['ED_status'] = None
                _temp_dict['Changed_sheets'] = \
                    original_dict['TDD_sets'][key]['Documents'][document][
                        'Sheets']
                _temp_dict['AMX_AM'] = \
                    original_dict['TDD_sets'][key]['Documents'][document][
                        'AMX'] if 'AMX' in \
                                  original_dict['TDD_sets'][key]['Documents'][
                                      document] else None
                _temp_dict['Descr_of_change'] = \
                    original_dict['TDD_sets'][key]['Documents'][document][
                        'Description']
                _temp_dict['New_rev?'] = None
                normalized_dict['TDD'].append(_temp_dict)

        for key in original_dict['SSC']:
            _temp_dict = {}
            _temp_dict['Type'] = 'Changed'
            _temp_dict['List_SSC'] = key
            _temp_dict['Name_KKS'] = None
            _temp_dict['Descr_of_change_KKS'] = None
            normalized_dict['SSC'].append(_temp_dict)

        for key in original_dict['Supp_descr_docs']:
            _temp_dict = {}
            _temp_dict['Type'] = 'Supporting'
            _temp_dict['File_name'] = key
            _temp_dict['File_content'] = original_dict['Supp_descr_docs'][key]
            normalized_dict['Support_files'].append(_temp_dict)

    return normalized_dict


def open_dict(obj, ini_str):
    for key in obj:
        if type(obj[key]) is dict:
            print(f"{ini_str * '   '}{key}:")
            open_dict(obj[key], ini_str + 1)
        else:
            print(f"{ini_str * '   '}{key}:   {obj[key]}")


def opener(dictus, spc, form='xml'):
    filler = '    '
    if form == 'xml':
        for key in dictus:
            if type(dictus[key]) != dict:
                print(f'{spc * filler}<{key}>')
                print(f'{(spc + 1) * filler}{dictus[key]}')
            else:
                print(f'{spc * filler}<{key}>')
                opener(dictus[key], spc + 1)
            print(f'{spc * filler}</{key}>')
    else:
        def js_ser(obj):
            if isinstance(obj, datetime.datetime):
                return obj.strftime('%d.%m.%Y')
            else:
                return None

        json_str = json.dumps(dictus, default=js_ser, ensure_ascii=False,
                              indent=2)
        return json_str


def from_diff_to_union_excel(normalized_dict):
    wb = smart_load_workbook('D:\\!Digital_twin\\!CR\\CR_parser/template.xlsx')
    ws = wb.worksheets[0]
    ws['B3'].value = normalized_dict['Change_request_No']
    ws['D3'].value = normalized_dict['Reg_date']
    ws['F3'].value = normalized_dict['Contr_change_coord']
    ws['B4'].value = normalized_dict['Type_of_changes']
    ws['D4'].value = normalized_dict['Constr_facility']
    ws['F4'].value = normalized_dict['Type_of_changed_doc']
    ws['H4'].value = normalized_dict['E-log?']
    ws['B6'].value = normalized_dict['Ini_org']
    ws['D6'].value = normalized_dict['Change_ini']
    ws['F6'].value = normalized_dict['Ini_internal_CR']
    ws['B8'].value = normalized_dict['Cod_of_reason']
    ws['D8'].value = normalized_dict['Other_reason']
    ws['B10'].value = normalized_dict['Descr_of_change']
    ws['B11'].value = normalized_dict['Rel_to_prev_CR']
    ws['B13'].value = normalized_dict['Approval_method']
    ws['D13'].value = normalized_dict['Just_simple']
    ws['B16'].value = normalized_dict['NSC_category']
    ws['D16'].value = normalized_dict['Impact_direct_123']
    ws['F16'].value = normalized_dict['Impact_DSA']
    ws['H16'].value = normalized_dict['Impact_structural_geom']
    ws['B17'].value = normalized_dict['Impact_nucl']
    ws['D17'].value = normalized_dict['Impact_fire']
    ws['F17'].value = normalized_dict['Impact_industrial']
    ws['H17'].value = normalized_dict['Impact_environment']
    ws['B18'].value = normalized_dict['Impact_TDD']
    ws['D18'].value = normalized_dict['Impact_lic_doc']
    ws['F18'].value = normalized_dict['Prompt_TDD?']
    ws['H18'].value = normalized_dict['Material_equivalent?']
    ws['B19'].value = normalized_dict['Comments_eng_eval']
    ws['B21'].value = normalized_dict['Impact_contract']
    ws['D21'].value = normalized_dict['Impact_cost']
    ws['F21'].value = normalized_dict['Impact_schedule']
    ws['B22'].value = normalized_dict['Comments_nont_ass']

    ws = wb.create_sheet('Signature list')
    ws.merge_cells('A1:G1')
    ws['A1'].value = 'CR / FCR signature list'
    ws['A2'].value = 'Role /\n Роль'
    ws['B2'].value = 'Position /\n Должность'
    ws['C2'].value = 'Name and Surname /\n Имя и Фамилия'
    ws['D2'].value = 'FMV number /\n Номер лицензии FMV'
    ws['E2'].value = 'HAEA registration /\n Регистрация HAEA'
    ws['F2'].value = 'Signature /\n Подпись'
    ws['G2'].value = 'Date /\n Дата'
    letter_massive = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    formating = excel_formating(ws)
    formating.sheet_view(1)
    formating.col_width(letter_massive, 34)
    formating.color([1, 1], 'ddebf7', 'ddebf7')
    _temp_list_of_cells = [[1, 1]] + [[2, x] for x in range(1, 8)]
    for cell in _temp_list_of_cells:
        formating.text_alignment(cell, 'center', 'center')
        if cell == [1, 1]:
            formating.text_font(cell, text_size=14)
        else:
            formating.text_font(cell, text_size=14, text_bold=True)

    _temp_counter = 2
    for key in normalized_dict['Signature_list']:
        _temp_counter += 1
        for letter in letter_massive:
            formating.text_alignment(f'{letter}{_temp_counter}', 'left',
                                     'center')
        ws[f'A{_temp_counter}'].value = key['Role']
        ws[f'B{_temp_counter}'].value = key['Position']
        ws[f'C{_temp_counter}'].value = key['Name_sur']
        ws[f'D{_temp_counter}'].value = key['FMV_number']
        ws[f'E{_temp_counter}'].value = key['HAEA_reg']
        ws[f'F{_temp_counter}'].value = None
        if isinstance(key['Date'], datetime.datetime):
            str_date = key['Date'].strftime('%d.%m.%Y')
        else:
            str_date = key['Date']
        ws[f'G{_temp_counter}'].value = str_date
    formating.borders('thin')

    ws = wb.create_sheet('4&6 Affected TDD')
    ws.merge_cells('A1:N1')
    ws['A1'].value = '4.1. Changed TDD'
    ws['A2'].value = 'Document set\ncode: /\nКод комплекта\nдокументов:'
    ws[
        'B2'].value = 'Document set\nname: /\nНаименование\nкомплекта\nдокументов:'
    ws[
        'C2'].value = 'Document set\nRevision: /\nРевизия\nкомплекта\nдокументов:'
    ws['D2'].value = 'Document set\nVersion: /\nВерсия комплекта\nдокументов:'
    ws['E2'].value = 'Document set\nStatus: /\nСтатус комплекта\nдокументов:'
    ws['F2'].value = 'ED set Code*: /\nКод выпускаемого\nдокумента*:'
    ws['G2'].value = 'ED set Name*: /\nНаименование\nвыпускаемого\nдокумента*:'
    ws['H2'].value = 'ED set Revision*: /\nРевизия\nвыпускаемого\nдокумента*:'
    ws['I2'].value = 'ED set Version*: /\nВерсия\nвыпускаемого\nдокумента*:'
    ws['J2'].value = 'ED set Status*: /\nСтатус\nвыпускаемого\nдокумента*:'
    ws['K2'].value = 'Changed sheets: /\nИзмененные\nлисты:'
    ws['L2'].value = 'Amx/Change\nversion: /\nНомер\nAM/Изменения:'
    ws['M2'].value = 'Change\ndescription: /\nОписание\nизменения:'
    ws[
        'N2'].value = 'New revision\nrequired: /\nТребуется ли\nвыпуск новой\nревизии:'

    letter_massive = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K',
                      'L', 'M', 'N']
    formating = excel_formating(ws)
    formating.sheet_view(1)
    formating.col_width(letter_massive, 17)
    formating.color([1, 1], 'ddebf7', 'ddebf7')
    _temp_list_of_cells = [[1, 1]] + [[2, x] for x in range(1, 15)]
    for cell in _temp_list_of_cells:
        if cell == [1, 1]:
            formating.text_font(cell, text_size=11)
            formating.text_alignment(cell, 'center', 'center')
        else:
            formating.text_font(cell, text_size=11, text_bold=True)
            formating.text_alignment(cell, 'left', 'center')

    _temp_counter = 2
    for key in normalized_dict['TDD']:
        if key['Type'] == 'TDD':
            _temp_counter += 1
            for letter in letter_massive:
                formating.text_alignment(f'{letter}{_temp_counter}', 'left',
                                         'center')
            ws[f'A{_temp_counter}'].value = key['Set_code']
            ws[f'B{_temp_counter}'].value = key['Set_name']
            ws[f'C{_temp_counter}'].value = key['Set_rev']
            ws[f'D{_temp_counter}'].value = key['Ser_version']
            ws[f'E{_temp_counter}'].value = key['Set_status']
            ws[f'F{_temp_counter}'].value = key['ED_code']
            ws[f'G{_temp_counter}'].value = key['ED_name']
            ws[f'H{_temp_counter}'].value = key['ED_rev']
            ws[f'I{_temp_counter}'].value = key['ED_version']
            ws[f'J{_temp_counter}'].value = key['ED_status']
            ws[f'K{_temp_counter}'].value = key['Changed_sheets']
            ws[f'L{_temp_counter}'].value = key['AMX_AM']
            ws[f'M{_temp_counter}'].value = key['Descr_of_change']
            ws[f'N{_temp_counter}'].value = key['New_rev?']

    _temp_counter += 1
    ws.merge_cells(f'A{_temp_counter}:N{_temp_counter}')
    ws[f'A{_temp_counter}'] = '6.3. Affected TDD'
    _temp_counter += 1
    for letter in letter_massive:
        ws[f'{letter}{_temp_counter}'].value = ws[f'{letter}2'].value
    formating.color([_temp_counter - 1, 1], 'ddebf7', 'ddebf7')
    _temp_list_of_cells = [[_temp_counter - 1, 1]] + [[_temp_counter, x] for x
                                                      in range(1, 15)]
    for cell in _temp_list_of_cells:
        if cell == [_temp_counter - 1, 1]:
            formating.text_font(cell, text_size=11)
            formating.text_alignment(cell, 'center', 'center')
        else:
            formating.text_font(cell, text_size=11, text_bold=True)
            formating.text_alignment(cell, 'left', 'center')

    for key in normalized_dict['TDD']:
        if key['Type'] == 'Configur':
            _temp_counter += 1
            for letter in letter_massive:
                formating.text_alignment(f'{letter}{_temp_counter}', 'left',
                                         'center')
            ws[f'A{_temp_counter}'].value = key['Set_code']
            ws[f'B{_temp_counter}'].value = key['Set_name']
            ws[f'C{_temp_counter}'].value = key['Set_rev']
            ws[f'D{_temp_counter}'].value = key['Ser_version']
            ws[f'E{_temp_counter}'].value = key['Set_status']
            ws[f'F{_temp_counter}'].value = key['ED_code']
            ws[f'G{_temp_counter}'].value = key['ED_name']
            ws[f'H{_temp_counter}'].value = key['ED_rev']
            ws[f'I{_temp_counter}'].value = key['ED_version']
            ws[f'J{_temp_counter}'].value = key['ED_status']
            ws[f'K{_temp_counter}'].value = key['Changed_sheets']
            ws[f'L{_temp_counter}'].value = key['AMX_AM']
            ws[f'M{_temp_counter}'].value = key['Descr_of_change']
            ws[f'N{_temp_counter}'].value = key['New_rev?']
    formating.borders('thin')

    ws = wb.create_sheet('4&6 Changed SSC')
    ws.merge_cells('A1:C1')
    ws['A1'].value = '4.2. Changed SSC'
    ws['A2'].value = 'List of affected SSC: /\nПеречень затронутых KKS:'
    ws['B2'].value = 'Name of affected KKS: / \nНаименование затронутого KKS:'
    ws['C2'].value = 'Description of KKS change: /\nОписание изменения KKS:'
    formating = excel_formating(ws)
    formating.sheet_view(1)
    letter_massive = ['A', 'B', 'C']
    formating.col_width(['A'], 30)
    formating.col_width(['B'], 47)
    formating.col_width(['C'], 54)
    formating.color([1, 1], 'ddebf7', 'ddebf7')
    _temp_list_of_cells = [[1, 1]] + [[2, x] for x in range(1, 4)]
    for cell in _temp_list_of_cells:
        if cell == [1, 1]:
            formating.text_font(cell, text_size=11)
            formating.text_alignment(cell, 'center', 'center')
        else:
            formating.text_font(cell, text_size=11, text_bold=True)
            formating.text_alignment(cell, 'left', 'center')

    _temp_counter = 2
    for key in normalized_dict['SSC']:
        if key['Type'] == 'Changed':
            _temp_counter += 1
            for letter in letter_massive:
                formating.text_alignment(f'{letter}{_temp_counter}', 'left',
                                         'center')
            ws[f'A{_temp_counter}'].value = key['List_SSC']
            ws[f'B{_temp_counter}'].value = key['Name_KKS']
            ws[f'C{_temp_counter}'].value = key['Descr_of_change_KKS']

    _temp_counter += 1
    ws.merge_cells(f'A{_temp_counter}:C{_temp_counter}')
    ws[f'A{_temp_counter}'] = '6.4. Impacted SSC'
    _temp_counter += 1
    for letter in letter_massive:
        ws[f'{letter}{_temp_counter}'].value = ws[f'{letter}2'].value
    formating.color([_temp_counter - 1, 1], 'ddebf7', 'ddebf7')
    _temp_list_of_cells = [[_temp_counter - 1, 1]] + [[_temp_counter, x] for x
                                                      in range(1, 4)]
    for cell in _temp_list_of_cells:
        if cell == [_temp_counter - 1, 1]:
            formating.text_font(cell, text_size=11)
            formating.text_alignment(cell, 'center', 'center')
        else:
            formating.text_font(cell, text_size=11, text_bold=True)
            formating.text_alignment(cell, 'left', 'center')

    for key in normalized_dict['SSC']:
        if key['Type'] == 'Configur':
            _temp_counter += 1
            for letter in letter_massive:
                formating.text_alignment(f'{letter}{_temp_counter}', 'left',
                                         'center')
            ws[f'A{_temp_counter}'].value = key['List_SSC']
            ws[f'B{_temp_counter}'].value = key['Name_KKS']
            ws[f'C{_temp_counter}'].value = key['Descr_of_change_KKS']
    formating.borders('thin')

    ws = wb.create_sheet('3&7 Supporting Files')
    ws.merge_cells('A1:B1')
    ws['A1'].value = '3.1. Supporting and describing documents'
    ws['A2'].value = 'File name: /\nИмя файла:'
    ws['B2'].value = 'File content: / \nСодержание файла:'
    formating = excel_formating(ws)
    formating.sheet_view(1)
    letter_massive = ['A', 'B']
    formating.col_width(['A'], 62)
    formating.col_width(['B'], 56)
    formating.color([1, 1], 'ddebf7', 'ddebf7')
    _temp_list_of_cells = [[1, 1]] + [[2, x] for x in range(1, 3)]
    for cell in _temp_list_of_cells:
        if cell == [1, 1]:
            formating.text_font(cell, text_size=11)
            formating.text_alignment(cell, 'center', 'center')
        else:
            formating.text_font(cell, text_size=11, text_bold=True)
            formating.text_alignment(cell, 'left', 'center')

    _temp_counter = 2
    for key in normalized_dict['Support_files']:
        if key['Type'] == 'Supporting':
            _temp_counter += 1
            for letter in letter_massive:
                formating.text_alignment(f'{letter}{_temp_counter}', 'left',
                                         'center')
            ws[f'A{_temp_counter}'].value = key['File_name']
            ws[f'B{_temp_counter}'].value = key['File_content']

    _temp_counter += 1
    ws.merge_cells(f'A{_temp_counter}:B{_temp_counter}')
    ws[f'A{_temp_counter}'] = '7.1. Documents justifying the decision'
    _temp_counter += 1
    for letter in letter_massive:
        ws[f'{letter}{_temp_counter}'].value = ws[f'{letter}2'].value
    formating.color([_temp_counter - 1, 1], 'ddebf7', 'ddebf7')
    _temp_list_of_cells = [[_temp_counter - 1, 1]] + [[_temp_counter, x] for x
                                                      in range(1, 3)]
    for cell in _temp_list_of_cells:
        if cell == [_temp_counter - 1, 1]:
            formating.text_font(cell, text_size=11)
            formating.text_alignment(cell, 'center', 'center')
        else:
            formating.text_font(cell, text_size=11, text_bold=True)
            formating.text_alignment(cell, 'left', 'center')

    for key in normalized_dict['Support_files']:
        if key['Type'] == 'Justifying':
            _temp_counter += 1
            for letter in letter_massive:
                formating.text_alignment(f'{letter}{_temp_counter}', 'left',
                                         'center')
            ws[f'A{_temp_counter}'].value = key['File_name']
            ws[f'B{_temp_counter}'].value = key['File_content']
    formating.borders('thin')

    file_name_list = normalized_dict['File_name'].split('.')
    file_name = '.'.join(file_name_list[:-1]) + '_NS.' + file_name_list[-1]
    output_dir = os.path.join(os.path.dirname(file_name), 'output')
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
    wb.save(os.path.join(output_dir, file_name.split('\\')[-1]))
    return normalized_dict['File_name']


class excel_formating:
    def __init__(self, ws):
        self.ws = ws

    def col_width(self, lett_list: list, value):
        for letter in lett_list:
            self.ws.column_dimensions[letter].width = value

    def text_alignment(self, place: list or str, hor, vert):
        if type(place) is list:
            cell = self.ws.cell(*place)
        else:
            cell = self.ws[place]
        cell.alignment = Alignment(horizontal=hor, vertical=vert,
                                   wrapText=True)

    def color(self, place: list, start_col: str, end_col: str):
        cell = self.ws.cell(*place)
        filling = PatternFill(start_color=start_col, end_color=end_col,
                              fill_type='solid')
        cell.fill = filling

    def text_font(self, place: list, text_size: int, text_bold=False):
        cell = self.ws.cell(*place)
        cell.font = Font(bold=text_bold, size=text_size)

    def sheet_view(self, page_number):
        self.ws.sheet_view.view = 'pageBreakPreview'
        self.ws.page_setup.fitToPage = True
        self.ws.page_setup.fitToWidth = page_number
        self.ws.page_setup.fitToHeight = False

    def borders(self, border_type_name):
        border_type = Border(
            left=Side(border_style=border_type_name, color='000000'),
            right=Side(border_style=border_type_name, color='000000'),
            top=Side(border_style=border_type_name, color='000000'),
            bottom=Side(border_style=border_type_name, color='000000'))
        for row in range(1, self.ws.max_row + 1):
            for col in range(1, self.ws.max_column + 1):
                self.ws.cell(row=row, column=col).border = border_type


def output(option):
    global result

    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))

    if option == 0:
        excel_path = 'D:\\!Digital_twin\\!CR\\CR_parser\\cr_test\\18_05\\'
    else:
        if option == 'pwd':
            excel_path = base_dir + '\\'
        else:
            user_input = input(
                'Введите ПОЛНЫЙ путь к папке с файлами (или нажмите Enter для текущей папки): ')
            if user_input.strip() == '':
                excel_path = base_dir + '\\'
            else:
                excel_path = user_input
                if excel_path and excel_path[-1] != '\\':
                    excel_path += '\\'

    print(f"\nИщем файлы в папке: {excel_path}")

    try:
        all_files = os.listdir(path=excel_path)
    except Exception as e:
        print(f"Ошибка при доступе к папке: {e}")
        return

    list_of_tables = list(filter(
        lambda x: (x.lower().endswith('.xls') or x.lower().endswith(
            '.xlsx')) and not x.startswith('~$'),
        all_files
    ))

    print(f"Найдено файлов Excel к обработке: {len(list_of_tables)}")
    if len(list_of_tables) == 0:
        print(
            "В указанной папке нет файлов формата .xls или .xlsx для обработки.")
        return

    result = {}

    for file in list_of_tables:
        full_path = f'{excel_path}{file}' if excel_path else file
        result[file] = main_func(full_path)

    if option == 0:
        for file in result:
            a = dicts_normalization(file)
            from_diff_to_union_excel(a)
    else:
        for dictus in result:
            if result.get(dictus):
                for code in result[dictus].get('TDD', {}):
                    if 'Impact' in result[dictus]['TDD'][code] and isinstance(
                            result[dictus]['TDD'][code]['Impact'], int):
                        bin_impact = f"{result[dictus]['TDD'][code]['Impact']:05b}"
                        bin_impact = [True if int(i) else False for i in
                                      bin_impact]
                        result[dictus]['TDD'][code]['Impact'] = {
                            'NS': bin_impact[0], 'FS': bin_impact[1],
                            'IS': bin_impact[2], 'ES': bin_impact[3],
                            'SS': bin_impact[4]}

                short_name = '.'.join(dictus.split('.')[:-1])
                file_name = excel_path + short_name + '.txt' if excel_path else short_name + '.txt'
                with open(file_name, 'wb+') as out_json:
                    out_json.write(b'\xff\xfe')
                    out_json.write(
                        opener(result[dictus], 0, form='json').encode(
                            'utf-16-le'))
                    print(
                        f'>>> Готово! Текстовый отчет сохранен: {short_name}.txt')


if __name__ == '__main__':
    try:
        output('pwd')
    except Exception as e:
        print(f"\nПроизошла критическая ошибка во время работы:")
        traceback.print_exc()
    finally:
        input("\nРабота завершена. Нажмите Enter, чтобы выйти...")